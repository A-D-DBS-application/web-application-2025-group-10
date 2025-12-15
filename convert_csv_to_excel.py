import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

INPUT_CSV = "klachten_export.csv"
OUTPUT_XLSX = "klachten_export_opgemaakt.xlsx"

def read_complaints_csv(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, sep=';', dtype=str)
    df = df.fillna('')
    bool_cols = [c for c in df.columns if df[c].str.lower().isin({'ja','nee'}).any()]
    for col in bool_cols:
        df[col] = df[col].map(lambda v: "Ja" if str(v).strip().lower() in ("ja","true","1") else "Nee")
    date_cols = [c for c in df.columns if "datum" in c.lower() or "date" in c.lower()]
    for col in date_cols:
        df[col] = pd.to_datetime(df[col], errors='coerce')
    return df, date_cols

def format_df_for_excel(df: pd.DataFrame, date_cols):
    for col in date_cols:
        if col in df.columns:
            mask = df[col].notna()
            df.loc[mask, col] = df.loc[mask, col].dt.strftime('%d-%m-%Y %H:%M').where(
                df.loc[mask, col].dt.strftime('%H:%M') != '00:00', df.loc[mask, col].dt.strftime('%d-%m-%Y')
            )
    return df

def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    if "Categorie type" in df.columns and "Status" in df.columns:
        summary = (
            df.groupby(["Categorie type", "Status"], dropna=False)
              .size()
              .reset_index(name="Aantal klachten")
              .sort_values(["Categorie type", "Status"])
        )
    else:
        summary = pd.DataFrame(columns=["Categorie type", "Status", "Aantal klachten"])
    return summary

def autofit_columns(ws):
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in col if cell.value not in (None, "")),
            default=0
        )
        max_length = max(max_length, 10)
        adjusted_width = min(max_length * 1.2 + 2, 60)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(adjusted_width, 15)
        for cell in col:
            cell.border = border
            if cell.row == 1:
                continue
            cell.alignment = Alignment(wrap_text=True, vertical="top")

def style_table(ws, table_name):
    table = Table(displayName=table_name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
    ws.freeze_panes = "A2"

def build_excel():
    df, date_cols = read_complaints_csv(INPUT_CSV)
    df = format_df_for_excel(df, date_cols)
    summary = build_summary(df)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Klachten", index=False)
        summary.to_excel(writer, sheet_name="Samenvatting", index=False)

    wb = load_workbook(OUTPUT_XLSX)
    ws = wb["Klachten"]
    summary_ws = wb["Samenvatting"]
    style_table(ws, "KlachtenTable")
    style_table(summary_ws, "SamenvattingTable")
    autofit_columns(ws)
    autofit_columns(summary_ws)
    wb.save(OUTPUT_XLSX)

if __name__ == "__main__":
    build_excel()
